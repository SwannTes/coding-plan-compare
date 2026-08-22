#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""血尿HCG检查脚本（统计分析 → 诊疗项目统计）

流程：
1. 关闭所有已打开的"诊疗项目统计"标签页（查询条件在服务端粘滞，
   旧面板里残留的筛选清不掉，必须开全新面板才是干净状态）
2. 点击"统计分析" → "诊疗项目统计"（模块 li id = CIC_module_IVC20，实测）
3. 开始/结束时间填入起止日期（命令行参数指定月份，如
   `python auto_fill_hcg_check.py 5` → 当年 5 月整月；留空默认本月 1 号至今天，
   与 fubao/monthly 脚本一致；也兼容 2026-05 这种 yyyy-mm 写法）
4. 循环项目名列表 ITEM_NAMES：清空并输入项目名 → 点"查询" →
   翻页爬取全部结果行 → 打印到控制台
5. 全部项目的结果导出到桌面一个 Excel（血尿HCG检查_起至止.xlsx），
   每个项目一个 sheet（sheet 名用项目名，非法字符替换、超长截断）

关键机制（沿用 auto_fill_monthly_check.py 实测趟出来的方案）：
- 系统的"导出"按钮不可用，改为直接读结果表格的 Ext store 数据（含全部字段）。
  本面板的分页工具栏不是 grid.getBottomToolbar()（那是个空 toolbar），
  而是面板内一个带 moveNext/cursor/pageSize 的独立组件，按"store 与 grid 相同"认出它；
  翻页必须走它的 moveNext（带游标校验），store.load({params:{start}}) 无效
- 系统会同时打开多个同名标签页（字段 id/name 冲突），
  一切操作限定在"当前活动标签页"的面板内进行（li.x-mytab-strip-active 的 id
  形如 标签条id__面板id）
- 查询按钮必须 Playwright 真实鼠标点击（JS 合成 click 无反应）：
  JS 定位后打 data-kimi-click="1" 标记，再 page.locator 点击；
  注意面板里还有 button.excel（导出）和 button.print（打印），只点 button.query
- 日期用 Ext API setValue 写入（直接改 DOM value 不会更新组件内部值）；
  本面板日期字段 name 是 KSSJ/JSSJ（不是就诊历史记录的 startDate/endDate），
  组件 id 是 ext-comp-* 不稳定的，按 name 定位
- 项目名称文本框 name=XMMC，用真实键盘输入（点击 → 全选 → 输入 → Tab），
  再回读 el.value 校验
- 项目名称匹配方式是"前缀匹配"（实测：输"尿妊娠"能匹配"尿妊娠试验-金标法"，
  输"绒毛膜"匹配不到"血清人绒毛膜促性腺激素测定-化学发光法"），
  所以血HCG要用前缀"血清人绒毛膜"，实际匹配到的项目名以结果 FYMC 为准
- 每次查询前在 store 上挂 load 钩子、点查询后等钩子触发，
  保证读到的是本次查询的新数据，不会拿到上一个项目的旧结果
- 结果 store 没有序号类唯一键，按整行 JSON 去重兜底
- 依赖 openpyxl 写 Excel（pip install --user openpyxl）
- 失败的步骤只警告不中断，脚本末尾统一汇总，方便人工补操作
"""

from playwright.sync_api import sync_playwright
from datetime import datetime
import calendar
import json
import os
import re
import sys
import time

FAILED_STEPS = []

DESKTOP = os.path.join(os.path.expanduser("~"), "Desktop")

# ===== 查询条件（按需修改） =====
# 月份由命令行参数指定（与 fubao/monthly 脚本一致，见 month_range 说明），此处只有项目名列表
# 项目名列表（前缀匹配，见文件头说明）：尿HCG + 血HCG
ITEM_NAMES = ["尿妊娠", "血清人绒毛膜"]

# 输出 Excel 的列：store 字段名 -> 中文表头（面板实测列）
FIELDS = [
    ("BRXM", "病人姓名"),
    ("FYMC", "项目名称"),
    ("YBFYBM", "医保费用编码"),
    ("YLDJ", "单价"),
    ("YLSL", "数量"),
    ("HJJE", "金额"),
    ("YSXM", "申请医生"),
    ("KDRQ", "开单日期"),
    ("SFRQ", "收费日期"),
    ("ZFRQ", "作废日期"),
]

# 当前活动标签页的面板定位 + 面板内组件查找（多标签页字段冲突，必须限定面板）
PANEL_JS = r"""
    const onScreen = el => {
        const r = el.getBoundingClientRect();
        if (r.width === 0 || r.height === 0) return false;
        if (r.x < -100 || r.y < -100) return false;
        let a = el;
        while (a && a !== document.body) {
            if (window.getComputedStyle(a).display === 'none') return false;
            a = a.parentElement;
        }
        return true;
    };
    const panelEl = () => {
        // 活动标签的 class 是 x-mytab-strip-active（有的版本是 x-mytab-strip-act）
        const tab = document.querySelector('li.x-mytab-strip-active, li.x-mytab-strip-act');
        if (!tab || tab.id.indexOf('__') < 0) return null;
        const panel = Ext.getCmp(tab.id.split('__')[1]);
        return panel && panel.el && panel.el.dom ? panel.el.dom : null;
    };
    // 结果表格：列模型里含 FYMC（项目名称）的 GridPanel（本面板本身就是 GridPanel）
    const findGrid = () => {
        const pel = panelEl();
        if (!pel) return null;
        let grid = null;
        Ext.ComponentMgr.all.each(c => {
            if (grid) return;
            if (c instanceof Ext.grid.GridPanel && c.getColumnModel && c.el && c.el.dom
                && pel.contains(c.el.dom)
                && (c.getColumnModel().config || []).some(col => col.dataIndex === 'FYMC')) {
                grid = c;
            }
        });
        return grid;
    };
    // 分页工具栏：getBottomToolbar() 拿到的是空 toolbar（实测踩过），
    // 真正的分页条是面板内带 moveNext 且 store 与结果表格相同的独立组件
    const findPager = grid => {
        const pel = panelEl();
        if (!pel || !grid) return null;
        let pager = null;
        Ext.ComponentMgr.all.each(c => {
            if (pager) return;
            if (typeof c.moveNext === 'function' && c.el && c.el.dom && pel.contains(c.el.dom)
                && c.store === grid.getStore()) {
                pager = c;
            }
        });
        return pager;
    };
"""

# 点查询前在 store 上挂 load 钩子（真实点击由 run_click 完成，evaluate 期间无法点击）
HOOK_SEARCH_JS = "() => {" + PANEL_JS + r"""
    const g = findGrid();
    if (!g) return false;
    window.__searchDone = false;
    g.getStore().on('load', () => { window.__searchDone = true; });
    return true;
}"""

# 翻页爬取全部结果：走分页工具栏 moveNext（游标校验），store.load({params:{start}}) 无效；
# store 没有序号类唯一键，按整行 JSON 去重兜底，同一行不会因翻页抖动重复入表
SCRAPE_JS = "async () => {" + PANEL_JS + r"""
    const grid = findGrid();
    if (!grid) return {error: '未找到结果表格组件'};
    const store = grid.getStore();
    const bbar = findPager(grid);
    if (!bbar) return {error: '未找到分页工具栏'};
    // 先回第一页（搜索后本来就在第一页，多退少补）
    if (bbar.cursor !== 0) {
        await new Promise(res => {
            const h = () => { store.un('load', h); res(); };
            store.on('load', h);
            bbar.moveFirst();
            setTimeout(res, 8000);
        });
    }
    const total = store.getTotalCount();
    const seen = {};
    let guard = 0;
    while (true) {
        store.each(rec => { seen[JSON.stringify(rec.data)] = rec.data; });
        if (bbar.cursor + bbar.pageSize >= total) break;
        const before = bbar.cursor;
        let advanced = false;
        for (let attempt = 0; attempt < 3 && !advanced; attempt++) {
            await new Promise(res => {
                const h = () => { store.un('load', h); res(); };
                store.on('load', h);
                bbar.moveNext();
                setTimeout(res, 8000);
            });
            advanced = bbar.cursor > before;
        }
        if (!advanced) return {error: '翻页失败(cursor=' + before + ')', collected: Object.keys(seen).length};
        if (++guard > 50) return {error: '翻页次数超限', collected: Object.keys(seen).length};
    }
    return {total, rows: Object.values(seen)};
}"""


def run_step(page, js, desc, timeout=10, quiet=False, record=True):
    """轮询执行JS（查找并操作，JS返回真值表示成功），直到成功或超时。
    成功后等待1秒，给系统反应时间。record=False 时超时不计入失败汇总（用于可重试的子步骤）。"""
    deadline = time.time() + timeout
    while True:
        try:
            ok = page.evaluate(js)
        except Exception:
            ok = False
        if ok:
            if not quiet:
                print(f"  [成功] {desc}")
            time.sleep(1)
            return True
        if time.time() >= deadline:
            print(f"  [超时] {desc} —— 未找到目标，请手动处理")
            if record:
                FAILED_STEPS.append(desc)
            return False
        time.sleep(0.3)


def run_click(page, locate_js, desc, verify_js=None, timeout=10, double=False, quiet=False, record=True):
    """轮询执行 locate_js 定位目标元素。locate_js 找到目标时给它打上临时标记
    data-kimi-click="1" 并返回 true，找不到返回 false。点击由 Playwright locator
    完成（真实鼠标事件 isTrusted=true，且自动滚动入视口、等待元素稳定、检测
    接收事件、失败自动重试）。
    提供 verify_js 时，点击后轮询 verify_js 确认生效，未生效会重新定位点击。
    record=False 时超时不计入失败汇总（用于可重试的子步骤）。"""
    # 每轮先清旧标记：上轮点击后元素可能被销毁重建，残留标记会导致 locator 点错
    clear_mark_js = ("() => document.querySelectorAll('[data-kimi-click]')"
                     ".forEach(e => e.removeAttribute('data-kimi-click'))")
    deadline = time.time() + timeout
    while True:
        try:
            page.evaluate(clear_mark_js)
            ok = page.evaluate(locate_js)
        except Exception:
            ok = False
        if ok:
            clicked = False
            try:
                target = page.locator('[data-kimi-click="1"]').first
                if double:
                    target.dblclick(timeout=3000)
                else:
                    target.click(timeout=3000)
                clicked = True
            except Exception:
                pass
            try:
                page.evaluate(clear_mark_js)
            except Exception:
                pass
            if not clicked:
                # locator 点击失败（超时/被遮挡/不可点）不能算成功，继续轮询重试
                pass
            elif verify_js is None:
                if not quiet:
                    print(f"  [成功] {desc}")
                time.sleep(1)
                return True
            else:
                time.sleep(0.5)
                try:
                    if page.evaluate(verify_js):
                        if not quiet:
                            print(f"  [成功] {desc}")
                        time.sleep(1)
                        return True
                except Exception:
                    pass
        if time.time() >= deadline:
            print(f"  [超时] {desc} —— 未找到目标，请手动处理")
            if record:
                FAILED_STEPS.append(desc)
            return False
        time.sleep(0.3)


def locate(body):
    """把定位语句包成 () => { ... } 函数，并注入 onScreen 助手。
    页面上大量隐藏窗口/模板副本（ExtJS 关闭是移到 -10000 而不是销毁），
    所有定位都要做"屏幕上可见"过滤。"""
    return "() => {" + r"""
    const onScreen = el => {
        const r = el.getBoundingClientRect();
        if (r.width === 0 || r.height === 0) return false;
        if (r.x < -100 || r.y < -100) return false;
        let a = el;
        while (a && a !== document.body) {
            if (window.getComputedStyle(a).display === 'none') return false;
            a = a.parentElement;
        }
        return true;
    };
""" + body + "\n}"


def search_and_wait(page, desc, timeout=30):
    """在当前活动面板内点查询按钮并等结果加载完成。
    查询按钮必须真实鼠标点击（合成 click 无反应），所以拆成：挂 load 钩子 → 真实点击 → 等钩子。
    等 load 钩子触发才返回，保证读到的是本次查询的新数据（多项目循环时不会读到旧结果）。
    注意面板里还有 button.excel（导出）和 button.print（打印），只认 button.query。"""
    if not run_step(page, HOOK_SEARCH_JS, desc + "-挂钩子", timeout=10, quiet=True):
        print(f"  [失败] {desc} —— 面板/表格未就绪")
        FAILED_STEPS.append(desc)
        return False
    if not run_click(page, "() => {" + PANEL_JS + r"""
        const pel = panelEl();
        if (!pel) return false;
        const btn = pel.querySelector('button.query');
        if (!btn || !onScreen(btn)) return false;
        btn.setAttribute('data-kimi-click', '1');
        return true;
    }""", desc, timeout=10, quiet=True):
        print(f"  [失败] {desc} —— 查询按钮点击失败")
        FAILED_STEPS.append(desc)
        return False
    if not run_step(page, "() => window.__searchDone === true", desc + "-加载", timeout=25, quiet=True):
        print(f"  [超时] {desc} —— 结果未加载，请手动处理")
        FAILED_STEPS.append(desc)
        return False
    print(f"  [成功] {desc}")
    return True


def type_field(page, field_name, text, desc):
    """真实键盘输入到面板内的输入框：点击 → 全选 → 输入 → Tab 提交。"""
    if not run_click(page, "() => {" + PANEL_JS + f"""
        const pel = panelEl();
        if (!pel) return false;
        const el = Array.from(pel.querySelectorAll('input[name={json.dumps(field_name)}]')).find(onScreen);
        if (!el) return false;
        el.setAttribute('data-kimi-click', '1');
        return true;
    }}""", desc + "-聚焦", timeout=10, quiet=True):
        print(f"  [失败] {desc} —— 输入框定位失败")
        FAILED_STEPS.append(desc)
        return False
    page.keyboard.press("Control+a")
    page.keyboard.type(text, delay=30)
    page.keyboard.press("Tab")
    time.sleep(0.5)
    ok = run_step(page, "() => {" + PANEL_JS + f"""
        const pel = panelEl();
        if (!pel) return false;
        const el = Array.from(pel.querySelectorAll('input[name={json.dumps(field_name)}]')).find(onScreen);
        return el ? el.value === {json.dumps(text)} : false;
    }}""", desc, timeout=5, quiet=True)
    if ok:
        print(f"  [成功] {desc}")
    else:
        print(f"  [失败] {desc}")
        FAILED_STEPS.append(desc)
    return ok


def scrape_all(page, desc):
    """翻页爬取当前查询结果的全部行，失败返回 None。"""
    try:
        r = page.evaluate(SCRAPE_JS)
    except Exception as e:
        print(f"  [失败] {desc}：爬取异常 {e}")
        FAILED_STEPS.append(desc)
        return None
    if not r or "rows" not in r:
        print(f"  [失败] {desc}：{r.get('error') if r else '无返回'}")
        FAILED_STEPS.append(desc)
        return None
    print(f"  [成功] {desc}：爬取 {len(r['rows'])} 行（总数 {r['total']}）")
    return r["rows"]


def safe_sheet_name(name, used):
    """把项目名变成合法的 Excel sheet 名：替换非法字符 \\ / ? * [ ] : ，
    最长 31 字符，重名时追加序号。"""
    s = re.sub(r'[\\/?*\[\]:]', "_", name).strip() or "结果"
    s = s[:31]
    base, i = s, 2
    while s in used:
        suffix = f"_{i}"
        s = base[:31 - len(suffix)] + suffix
        i += 1
    used.add(s)
    return s


def save_excel(results, filename, desc):
    """把各项目的爬取结果写成多 sheet 的 xlsx 存到桌面。
    results: [(项目名, 行列表或None), ...]，None 表示该项目爬取失败，跳过。"""
    try:
        from openpyxl import Workbook
        wb = Workbook()
        wb.remove(wb.active)  # 删掉默认空 sheet，全部按项目建
        used = set()
        for item_name, rows in results:
            if rows is None:
                continue
            ws = wb.create_sheet(safe_sheet_name(item_name, used))
            ws.append([h for _, h in FIELDS])
            for row in rows:
                ws.append([row.get(k) if row.get(k) is not None else "" for k, _ in FIELDS])
        if not wb.sheetnames:
            print(f"  [失败] {desc}：没有任何项目的数据可写")
            FAILED_STEPS.append(desc + "(写Excel)")
            return
        path = os.path.join(DESKTOP, filename)
        wb.save(path)
        summary = "、".join(f"{n}{len(r)}条" for n, r in results if r is not None)
        print(f"  [成功] {desc}：{summary} → {path}")
    except Exception as e:
        print(f"  [失败] {desc}：写 Excel 出错 {e}")
        FAILED_STEPS.append(desc + "(写Excel)")


def print_rows(rows):
    """把结果行打印到控制台（表头 + 每行按字段排列）。"""
    if not rows:
        print("  （无数据行）")
        return
    headers = [h for _, h in FIELDS]
    print("  " + " | ".join(headers))
    print("  " + "-" * 60)
    for row in rows:
        cells = [str(row.get(k) if row.get(k) is not None else "") for k, _ in FIELDS]
        print("  " + " | ".join(cells))


def month_range(month_arg=""):
    """返回起止日期 ('yyyy-mm-01', 'yyyy-mm-月末')。
    与 fubao/monthly 脚本一致：month_arg 为空 → 本月 1 号到今天；为 1-12 → 该年该月整月，
    月份大于当前月份时取上一年（如 1 月查去年 12 月）。
    另外兼容 'yyyy-mm' 写法（如 2026-05），方便直接指定年月。
    月末日期由 calendar.monthrange 算，28/29/30/31 天自动处理。"""
    now = datetime.now()
    if month_arg:
        if re.fullmatch(r"\d{4}-\d{1,2}", month_arg):  # yyyy-mm 写法
            year, month = map(int, month_arg.split("-"))
            if not 1 <= month <= 12:
                raise ValueError(f"月份必须是 1-12，收到: {month_arg!r}")
        else:
            month = int(month_arg)
            if not 1 <= month <= 12:
                raise ValueError(f"月份必须是 1-12，收到: {month_arg!r}")
            year = now.year - 1 if month > now.month else now.year
        last_day = calendar.monthrange(year, month)[1]
        end = f"{year}-{month:02d}-{last_day:02d}"
    else:
        year, month = now.year, now.month
        end = now.strftime("%Y-%m-%d")
    return f"{year}-{month:02d}-01", end


def hcg_check():
    print("1. 开始启动...")
    # 月份参数：命令行第 1 个参数（面板输入框也是传到这里），留空默认本月至今
    month_arg = sys.argv[1].strip() if len(sys.argv) > 1 else ""
    try:
        start_date, end_date = month_range(month_arg)
    except ValueError as e:
        print(f"参数错误：{e}")
        sys.exit(1)
    print(f"   查询月份: {start_date[:7]}（{start_date} 至 {end_date}）")
    with sync_playwright() as p:
        print("2. 尝试连接Chrome...")
        try:
            browser = p.chromium.connect_over_cdp("http://localhost:9222")
        except Exception:
            print("错误：连接失败。请先运行 启动调试Chrome并打开网址.py 并登录系统。")
            return
        print("3. 连接成功!")

        target_url = "172.17.8.14:8780"
        page = None
        for ctx in browser.contexts:
            for pg in ctx.pages:
                if target_url in pg.url:
                    page = pg
                    break
            if page:
                break
        if not page:
            print(f"错误：未找到包含 {target_url} 的页面")
            return
        print(f"4. 当前页面标题: {page.title()}")

        # ========== 1. 关闭所有已打开的"诊疗项目统计"标签页 ==========
        # 查询条件在服务端/会话层面粘滞，旧面板残留的筛选用空值清不掉，
        # 必须开全新面板才是干净状态。没有旧标签是正常情况，直接探测不报错
        print("5. 关闭旧的诊疗项目统计标签页...")
        closed = 0
        LOCATE_CLOSE = locate(r"""
            const tab = Array.from(document.querySelectorAll('li.x-mytab-strip-closable'))
                .find(li => (li.textContent || '').includes('诊疗项目统计') && onScreen(li));
            if (!tab) return false;
            const close = tab.querySelector('a.x-mytab-strip-close');
            if (!close) return false;
            close.setAttribute('data-kimi-click', '1');
            return true;
        """)
        for _ in range(5):  # 最多关 5 个，防止异常死循环
            try:
                found = page.evaluate(LOCATE_CLOSE)
            except Exception:
                found = False
            if not found:
                break
            try:
                page.locator('[data-kimi-click="1"]').first.click(timeout=3000)
                closed += 1
            except Exception:
                break
            finally:
                page.evaluate("() => document.querySelectorAll('[data-kimi-click]')"
                              ".forEach(e => e.removeAttribute('data-kimi-click'))")
            time.sleep(1)
        if closed:
            print(f"  [成功] 关闭了 {closed} 个旧标签页")
        else:
            print("  （没有旧标签页，跳过）")

        # ========== 2. 点击统计分析 → 诊疗项目统计（全新面板） ==========
        print("6. 点击统计分析...")
        run_click(page, locate(r"""
            const links = Array.from(document.querySelectorAll('a')).filter(a =>
                (a.textContent || '').trim() === '统计分析' && onScreen(a));
            if (!links.length) return false;
            links[0].setAttribute('data-kimi-click', '1');
            return true;
        """), "点击统计分析", timeout=15)

        print("7. 点击诊疗项目统计...")
        # 菜单项 LI id 为 CIC_module_IVC20（模块编码实测稳定）；兜底按文本匹配，
        # 但要排除已打开的标签页（x-mytab）里同名的那个
        run_click(page, locate(r"""
            let link = null;
            const li = document.getElementById('CIC_module_IVC20');
            if (li) link = li.querySelector('a') || li;
            if (!link || !onScreen(link)) {
                link = Array.from(document.querySelectorAll('a')).find(a =>
                    (a.textContent || '').trim() === '诊疗项目统计' && onScreen(a)
                    && !a.closest('.x-mytab-strip, [class*="x-mytab"]'));
            }
            if (!link || !onScreen(link)) return false;
            link.setAttribute('data-kimi-click', '1');
            return true;
        """), "点击诊疗项目统计", timeout=15)
        # 面板加载较慢，多等一会
        time.sleep(2)

        # ========== 3. 填写日期范围（整个循环共用，只填一次） ==========
        print(f"8. 填写时间范围: {start_date} 至 {end_date} ...")
        # 日期必须用 Ext API setValue 写入（直接改 DOM value 不会更新组件内部值，
        # 查询会拿旧日期）；本面板字段 name 是 KSSJ/JSSJ，组件 id 是 ext-comp-*
        # 不稳定的，按 name 定位后取 el.id 找组件
        run_step(page, "() => {" + PANEL_JS + f"""
            const pel = panelEl();
            if (!pel) return false;
            const setDate = (name, val) => {{
                const el = Array.from(pel.querySelectorAll('input[name="' + name + '"]')).find(onScreen);
                if (!el) return false;
                const c = Ext.getCmp(el.id);
                if (c && c.setValue) {{ c.setValue(val); return c.getRawValue() === val; }}
                el.value = val;
                el.dispatchEvent(new Event('change', {{ bubbles: true }}));
                return el.value === val;
            }};
            return setDate('KSSJ', {json.dumps(start_date)})
                && setDate('JSSJ', {json.dumps(end_date)});
        }}""", "填写时间范围")

        # ========== 4. 循环每个项目名：输入 → 查询 → 爬取 ==========
        results = []  # [(项目名, 行列表或None)]
        for idx, item_name in enumerate(ITEM_NAMES, 1):
            print(f"9.{idx} 项目「{item_name}」...")
            if not type_field(page, "XMMC", item_name, f"项目名称={item_name}"):
                results.append((item_name, None))
                continue
            # search_and_wait 内部等 store load 钩子触发才返回，
            # 保证爬到的是本项目的新数据，不是上一个项目的旧结果
            if search_and_wait(page, f"查询 {item_name}"):
                rows = scrape_all(page, f"爬取 {item_name}")
            else:
                rows = None
            results.append((item_name, rows))
            if rows is not None:
                matched = sorted({str(r.get("FYMC") or "") for r in rows})
                print(f"  [明细] 「{item_name}」共 {len(rows)} 行，"
                      f"实际匹配项目名：{'、'.join(matched) if matched else '（无）'}")
                if rows:
                    print_rows(rows)
                else:
                    # 0 行不一定是脚本问题，明确提示人工核对条件
                    print("  [提示] 查询结果 0 行：条件和面板填写均已校验成功，"
                          "请人工确认该时间段内确实无此项目数据")

        # ========== 5. 导出 Excel（一个文件，每项目一个 sheet） ==========
        print("10. 导出 Excel...")
        save_excel(results, f"血尿HCG检查_{start_date}至{end_date}.xlsx", "血尿HCG检查结果")

        # ========== 汇总 ==========
        print("=" * 40)
        if FAILED_STEPS:
            print(f"完成，但有 {len(FAILED_STEPS)} 个步骤未成功，请手动检查：")
            for s in FAILED_STEPS:
                print(f"  - {s}")
        else:
            print("全部步骤执行成功！")


if __name__ == "__main__":
    hcg_check()
